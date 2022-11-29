from typing import Union, List, Tuple

from fastapi import FastAPI, File, UploadFile, Form, Body, Depends, Response
from fastapi.responses import StreamingResponse


from pydantic import BaseModel, parse_obj_as # for資料驗證

from fastapi.middleware.cors import CORSMiddleware # CORS


# for 隨機抽取
import random

# for 輸出excel座位表
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
import os
import zipfile
from io import StringIO, BytesIO

import csv
import codecs
import json


app = FastAPI()

# 設置CORS
origins = [
    "http://localhost",
    "http://localhost:8080",
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def read_root():
    return {"Hello": "World"}


@app.post("/submit", response_class = StreamingResponse)
def submit(classRooms: str = Form(), file: UploadFile = File(...)):

    # 解析csv檔 -> 取得 studentDict
    csvReader = csv.DictReader(codecs.iterdecode(file.file, 'utf-8'))
    studentDict = {}
    for rows in csvReader:             
        key = rows['id']  # Assuming a column named 'id' to be the primary key
        studentDict[key] = rows  
    
    # 解析classRooms -> 取得 classRoomList
    print(classRooms)
    classRooms = json.loads(classRooms)

    # 1. 學生隨機分班
    myStudentList = list(studentDict.items())
    random.shuffle(myStudentList)
    studentList = []
    for classRoom in classRooms:
        totalNum = int(classRoom['finalSeat'])
        studentList.append(myStudentList[:totalNum])
        myStudentList = myStudentList[totalNum:]

    # 2. 安排座位
    sheetCol = {0: 'A', 1: 'B', 2: 'C', 3: 'D', 4: 'E', 5: 'F', 6: 'G', 7: 'H', 8: 'I', 9: 'J', 10: 'K', 11: 'L', 12: 'M', 13: 'N', 14: 'O', 15: 'P', 16: 'Q', 17: 'R', 18: 'S', 19: 'T', 20: 'U' }
    for classRoom, students in zip(classRooms, studentList):
        seats = classRoom['diagram']
        for seat in seats:
            seat['row']+=1
            seat['col'] = sheetCol[seat['col']]
            if(seat['status']==1):
                students_pop = students.pop()
                seat['name'] = list(students_pop[1].values())[0]
                seat['id'] = list(students_pop[1].values())[1]
            else:
                seat['name'] = None
                seat['id'] = None
            if(len(students)<0):
                break
    # 3. 畫圖
    fileNames = []
    for classRoom in classRooms:
        fileName = classRoom['roomNo'] + '.xlsx'

        book = Workbook()
        sheet = book.active
        
        for seat in classRoom['diagram']:
            insertIndex = seat['col'] + str(seat['row'])
            if seat['status'] == -1 :
                insertData = '"' + str(seat['row']) + '"'
                sheet[insertIndex].value = '= %s' % insertData
                sheet[insertIndex].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                sheet.row_dimensions[seat['row']].height =  60.0
                sheet.column_dimensions[seat['col']].width = 7.5
                sheet[insertIndex].fill = PatternFill("solid", fgColor="C4C4C4")
                sheet[insertIndex].border = Border(left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'))
            else:
                if seat['status'] == 1 :
                    insertData = '"' + seat['name'] + '"' + "&CHAR(10)&" + '"' + seat['id'] + '"'
                else:
                    insertData = ""
                sheet[insertIndex].value = '= %s' % insertData
                sheet[insertIndex].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
                sheet.row_dimensions[seat['row']].height =  60.0
                sheet.column_dimensions[seat['col']].width = 15.0
                sheet[insertIndex].border = Border(left=Side(border_style='thin',color='000000'),right=Side(border_style='thin',color='000000'),top=Side(border_style='thin',color='000000'),bottom=Side(border_style='thin',color='000000'))

        book.save(fileName)
        fileNames.append(fileName)
    
    return zipfiles(fileNames)


@app.post("/test")
def image_filter():
    file_list = ['I1-017.xlsx', 'I1-105.xlsx']
    return zipfiles(file_list)

def zipfiles(filenames):
    zip_subdir = "archive"
    zip_filename = "%s.zip" % zip_subdir

    # Open StringIO to grab in-memory ZIP contents
    s = BytesIO()
    # The zip compressor
    zf = zipfile.ZipFile(s, "w")

    for fpath in filenames:
        # Calculate path for file in zip
        fdir, fname = os.path.split(fpath)
        zip_path = os.path.join(zip_subdir, fname)

        # Add file, at correct path
        zf.write(fpath, zip_path)

    # Must close zip for all contents to be written
    zf.close()
    print(zf.namelist())
    # Grab ZIP file from in-memory, make response with correct MIME-type
    resp = StreamingResponse(iter([s.getvalue()]), media_type="application/x-zip-compressed", headers={
        'Content-Disposition': f'attachment;filename={zip_filename}'
    })
    return resp

